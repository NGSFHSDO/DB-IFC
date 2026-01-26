#!/usr/bin/env python
# -*- coding: utf-8 -*-
from __future__ import annotations

import time
from pathlib import Path
from typing import List, Tuple

import pandas as pd
from langchain_openai import OpenAIEmbeddings
from dotenv import load_dotenv
from tqdm import tqdm
from openpyxl import Workbook, load_workbook

DEFAULT_INPUT_FILE = "data\stockAnalysisReport_2025-10-13_2025-12-31.parquet"
DEFAULT_OUTPUT_XLSX = "data/kospidaq_embeddings_0.xlsx"
DEFAULT_MAX_ROWS = None  # Set to None to process all rows
DEFAULT_SAVE_EVERY = 100000
DEFAULT_START_ROW = 0


def _clean_excel_text(text: str) -> str:
    # Excel rejects control chars (0x00-0x1F) except tab/newline.
    return "".join(ch for ch in text if ch == "\t" or ch == "\n" or ch == "\r" or ord(ch) >= 32)


def normalize_reports(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    if "date" in df.columns:
        df["date"] = pd.to_datetime(df["date"], errors="coerce")
    for col in ("broker", "content", "industry", "ticker", "equity", "title"):
        if col in df.columns:
            df[col] = df[col].fillna("").astype(str)
    return df


def _records_to_rows(records: List[dict], embedding_dim: int) -> List[List[object]]:
    embedding_cols = [f"embedding_{i}" for i in range(1, embedding_dim + 1)]
    columns = ["date", "ticker", "broker", "target_price", "rating", *embedding_cols]
    rows: List[List[object]] = []
    for record in records:
        rows.append([record.get(col, "") for col in columns])
    return rows


def _append_excel(records: List[dict], embedding_dim: int, output_path: Path) -> None:
    if not records or not embedding_dim:
        return
    embedding_cols = [f"embedding_{i}" for i in range(1, embedding_dim + 1)]
    columns = ["date", "ticker", "broker", "target_price", "rating", *embedding_cols]
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if output_path.exists():
        wb = load_workbook(output_path)
        ws = wb.active
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        if header != columns:
            raise ValueError("Excel header does not match expected columns; cannot append safely.")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(columns)
    for row in _records_to_rows(records, embedding_dim):
        ws.append(row)
    wb.save(output_path)


def _infer_embedding_dim_from_header(header: List[object]) -> int:
    max_idx = 0
    for col in header:
        if not isinstance(col, str) or not col.startswith("embedding_"):
            continue
        suffix = col.split("_", 1)[-1]
        if suffix.isdigit():
            max_idx = max(max_idx, int(suffix))
    return max_idx


def _read_existing_excel_state(output_path: Path) -> Tuple[int, int]:
    if not output_path.exists():
        return 0, 0
    wb = load_workbook(output_path, read_only=True)
    ws = wb.active
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    embedding_dim = _infer_embedding_dim_from_header(header)
    if not embedding_dim:
        raise ValueError("Could not infer embedding dimension from existing Excel header.")
    processed_count = max(ws.max_row - 1, 0)
    return processed_count, embedding_dim


def build_embedding_records(
    df: pd.DataFrame,
    max_rows: int | None,
    save_every: int,
    output_path: Path,
    existing_embedding_dim: int,
) -> Tuple[List[dict], int, int]:
    embedder = OpenAIEmbeddings()
    records: List[dict] = []
    total = len(df)
    embedding_dim = existing_embedding_dim
    written_count = 0
    for i, (_, row) in enumerate(tqdm(df.iterrows(), total=total, desc="Embedding", unit="row"), start=1):
        broker = row.get("broker", "")
        ticker = row.get("ticker", "")
        date_val = row.get("date")
        date_str = ""
        if pd.notna(date_val):
            date_str = pd.Timestamp(date_val).date().isoformat()
        target_price = row.get("target_price", "")
        rating = row.get("rating", "")
        content = row.get("content", row.get("body", ""))
        text = _clean_excel_text(str(content).strip())
        vector = embedder.embed_documents([text])[0]
        if not embedding_dim:
            embedding_dim = len(vector)
        embedding_cols = {f"embedding_{j+1}": float(v) for j, v in enumerate(vector)}
        records.append(
            {
                "date": date_str,
                "ticker": str(ticker),
                "broker": broker,
                "target_price": str(target_price),
                "rating": str(rating),
                **embedding_cols,
            }
        )
        if i % 100 == 0 or i == total:
            print(f"Embedded {i}/{total} documents ({i/total:.1%})")
        if save_every and i % save_every == 0:
            _append_excel(records, embedding_dim, output_path)
            print(f"Saved {i} rows to {output_path}")
            written_count += len(records)
            records = []
        if max_rows and i >= max_rows:
            break
    return records, embedding_dim, written_count


def main() -> None:
    load_dotenv()
    input_path = Path(DEFAULT_INPUT_FILE).resolve()
    output_path = Path(DEFAULT_OUTPUT_XLSX).resolve()
    max_rows = DEFAULT_MAX_ROWS if DEFAULT_MAX_ROWS is None else int(DEFAULT_MAX_ROWS)
    save_every = int(DEFAULT_SAVE_EVERY)
    start_row = int(DEFAULT_START_ROW)

    if not input_path.exists():
        raise FileNotFoundError(f"Parquet file not found: {input_path}")

    df = normalize_reports(pd.read_parquet(input_path))
    if df.empty:
        raise RuntimeError("No rows found in input parquet file.")

    embedding_dim = 0
    processed_count = 0
    if output_path.exists():
        try:
            processed_count, embedding_dim = _read_existing_excel_state(output_path)
        except Exception:
            backup_path = output_path.with_name(
                f"{output_path.stem}_corrupt_{time.strftime('%Y%m%d_%H%M%S')}{output_path.suffix}"
            )
            output_path.replace(backup_path)
            print(f"Corrupt Excel moved to {backup_path}")
            processed_count = 0
            embedding_dim = 0
        else:
            if processed_count >= len(df):
                print(f"All {processed_count} rows already embedded. Nothing to do.")
                return
            df = df.iloc[processed_count:].reset_index(drop=True)
    elif start_row:
        if start_row >= len(df):
            raise RuntimeError(f"DEFAULT_START_ROW={start_row} is beyond input rows ({len(df)}).")
        df = df.iloc[start_row:].reset_index(drop=True)
        processed_count = start_row

    records, embedding_dim, written_count = build_embedding_records(
        df, max_rows, save_every, output_path, embedding_dim
    )
    _append_excel(records, embedding_dim, output_path)
    written_count += len(records)
    print(f"Wrote {processed_count + written_count} rows to {output_path}")


if __name__ == "__main__":
    main()
